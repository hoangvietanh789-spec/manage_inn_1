"""
Created on Mon Nov 17 14:58:00 2025

@author: admin
"""
# import os
# print(os.getcwd())

# com = 'E:/SETUP/Task_UAT/invest' if os.getcwd() == 'E:\SETUP\Task_UAT\invest' else ''
# db_file = com + "/content/drive/MyDrive/Dau_tu/data/inn.db"
# file_price = com + "/content/drive/MyDrive/Dau_tu/data/prices.json"
# file_room = com + "/content/drive/MyDrive/Dau_tu/data/rooms.json"
# file_tenant = com + "/content/drive/MyDrive/Dau_tu/data/tenants.json"
# file_account = com + "/content/drive/MyDrive/Dau_tu/data/accounts.json"
# file_cashbanoi = com + "/content/drive/MyDrive/Dau_tu/report/cash_banoi.xlsx"
# file_report = com + "/content/drive/MyDrive/Dau_tu/report/rent_report.xlsx"  
# file_cashflow = com + "/content/drive/MyDrive/Dau_tu/report/cash_flow.xlsx"  
# file_hangthang = com + "/content/drive/MyDrive/Dau_tu/report/Hang thang.xlsx"

# =============================================================================
# mount drive folder
# =============================================================================
def safe_mount_drive(mount_point="/content/drive"):
    # if com  == 'E:/SETUP/Task_UAT/invest':
    #     return
    import os
    import io
    import contextlib
    from google.colab import drive
    if not os.path.ismount(mount_point):
        f = io.StringIO()
        with contextlib.redirect_stdout(f):
            drive.mount(mount_point)

db_file = "/content/drive/MyDrive/Dau_tu/data/inn.db"
file_price = "/content/drive/MyDrive/Dau_tu/data/prices.json"
file_room = "/content/drive/MyDrive/Dau_tu/data/rooms.json"
file_tenant = "/content/drive/MyDrive/Dau_tu/data/tenants.json"
file_account = "/content/drive/MyDrive/Dau_tu/data/accounts.json"
file_cashbanoi = "/content/drive/MyDrive/Dau_tu/report/cash_banoi.xlsx"
file_report = "/content/drive/MyDrive/Dau_tu/report/rent_report.xlsx"  
file_cashflow = "/content/drive/MyDrive/Dau_tu/report/cash_flow.xlsx"  
file_hangthang = "/content/drive/MyDrive/Dau_tu/report/Hang thang.xlsx"


def brief():
    print("""
        ====================================
            tong_diennuoc("{this_month}", sodien, tien dien, sonuoc, tiennuoc): Nhập số tiền điện, tiền nước tổng theo tháng
            tinhtien(): Tính toán số tiền phải thanh toán của phòng theo tháng nhập/ tất cả
            doanhthu(): Tính toán cân đối thu chi
            chikhac(): ghi nhận giao dịch vào tk CTG và tạo thông tin vào bảng qua chikhac1()
                chikhac1(date, noidung_chi, sotien_chi, ghichu)
            xoa_chi_khac(record_id): xóa giao dịch trên CTG và cập nhật bảng chikhac
        ====================================
            pay(): Nhập số tiền thanh toán theo từng phòng
            unpay(): Xóa số tiền đã trả của 01 phòng và xóa luôn giao dịch được mark trong tk CTG
            diennuoc(): Nhập số CÔNG TƠ điện nước theo từng phòng
        ====================================
            query(): trả json các bảng rooms, tenants, prices
            querydf(): trả df các bảng cashflow, tong_diennuoc, chikhac
        ====================================
            view(): Mở link web
        ====================================
            new_month(): khởi động tháng mới
            add_room(room_data): thêm 01 room mới gối đầu giữa tháng
            add_tenant(tenant_data): thêm thông tin người dùng mới
            automap_tenant(): tự động cập nhật từ tenants vào rooms
            manualmap_tenant(): thủ công cập nhật từ tenants vào rooms
            change_tenant_status(**kwargs): cập nhật người trả phòng
            reset_room(*room_reset): xóa dữ liệu room trả phòng
        ====================================
            update: cập nhật json
            import_json(): nhập lại dữ liệu từ các file json
        
""")


            
# =============================================================================
# Tính toán số tiền phải thanh toán của phòng theo tháng nhập/ tất cả
# =============================================================================
def tinhtien(*month_input):
    import json
    import pandas as pd
    from datetime import datetime
    
    today = datetime.now()
    this_month = datetime.strftime(today, "%Y%m")

    safe_mount_drive()
    tong_diennuoc = querydf('tong_diennuoc')
    if this_month not in list(tong_diennuoc['Month']):
        print("Chưa có hóa đơn tổng điện nước tháng này")
        return

    tenant = query('tenants')
    account = query('accounts')

    price = query('prices')
    electric_price = price[this_month]['electric_price']
    water_price = price[this_month]['water_price']

    data = query('rooms')
    all_records = []
    print(data.keys())
    month_tocal = [this_month] 
    if len(month_input) == 0:
        ask = input("Month to calculate [add / all]: ")
        if ask == 'all':
            month_tocal = list(data.keys())
        else:
            while ask !=  '':
                if ask in data.keys():
                    month_tocal.append(ask)
                else:
                    print(ask, "not in data")
                ask = input(" ")
            for i in month_tocal:
                try:
                    datetime.strptime(i, "%Y%m")
                except Exception as ex:
                    print(ex)
    month_tocal = list(set(month_tocal))

    def calculate(room,info):
        if info["electric_start"] is not None and info["electric_end"] is not None:
            electric_fee = (info["electric_end"] - info["electric_start"]) * electric_price
        else:
            electric_fee = 0

        if info["water_start"] is not None and info["water_end"] is not None:
            water_fee = (info["water_end"] - info["water_start"]) * water_price
        else:
            water_fee = 0

        bill = info["rent_price"] or 0
        payment = info["payment"] or 0
        if this_month in info['prepayment']:           
            prepayment = info['prepayment'][this_month] or 0
        else:
            prepayment = 0
        bill = bill + electric_fee if electric_fee >= 0 else bill
        bill = bill +  water_fee if water_fee >= 0 else bill
        if bill <= prepayment:
            due_amount = 0
        elif bill > prepayment:
            due_amount = bill - prepayment - payment if bill - prepayment - payment > 0 else 0

        # Cập nhật vào dict
        info["electric_fee"] = electric_fee
        info["water_fee"] = water_fee
        info["bill"] = bill
        info["due_amount"] = due_amount

        # Auto update status
        if info["start_date"]:
            info["status"] = "rented"
        else:
            info["status"] = "available"

        return(info)

    for month, rooms in data.items():
        for room, info in rooms.items():
            if month in month_tocal and info['status'] == 'rented':
                info = calculate(room, info)
                for fo in ['electric_fee','water_fee','bill','due_amount']: # chỉ cập nhật các trường tính toán riêng prepay sẽ cập nhật khi new_month
                     update('rooms', f'{month}.{room}.{fo}', info[fo]) #update trước rồi mới convert để gen file như dưới đây
            info['bill'] = 0 if info['bill'] == None else info['bill']
            info['rent_price'] = 0 if info['rent_price'] == None else info['rent_price']
            info['electric_fee'] = 0 if info['electric_fee'] == None else info['electric_fee']
            info['electric_end'] = 0 if info['electric_end'] == None else info['electric_end']
            info['electric_start'] = 0 if info['electric_start'] == None else info['electric_start']
            info['water_fee'] = 0 if info['water_fee'] == None else info['water_fee']
            info['water_end'] = 0 if info['water_end'] == None else info['water_end']
            info['water_start'] = 0 if info['water_start'] == None else info['water_start']
            info['payment'] = 0 if info['payment'] == None else info['payment']
            if isinstance(info['prepayment'], dict) and this_month in info['prepayment']:           
                prepayment = info['prepayment'][this_month] or 0
            else:
                prepayment = 0
           
            all_records.append({
                "month": month,
                "room": room,
                "tenant": info["phone"],
                "name": info["name"],
                "rent_price": info["rent_price"],
                "electric_fee": info["electric_fee"],
                "water_fee": info["water_fee"],
                "bill": info["bill"],
                "prepayment": prepayment,
                "payment": info["payment"],
                "due_amount": info["due_amount"],
                "status": info["status"],
                "zalo_link": f"https://zalo.me/{info['phone']}" if info.get("phone") else "",
                "notify":f"""Tổng: {info["bill"]:,.0f}, trong đó:
    - Tiền thuê: {info["rent_price"]:,.0f}
    - Tiền điện: {info["electric_fee"]:,.0f} = ({info["electric_end"]:,.0f} - {info["electric_start"]:,.0f}) * {electric_price:,.0f}đ/kWh
    - Tiền nước: {info["water_fee"]:,.0f} = ({info["water_end"]:,.0f} - {info["water_start"]:,.0f}) * {water_price:,.0f}đ/m3
    - Đã thanh toán/trả trước: {info["payment"]:,.0f}
    - Còn thiếu: {info["due_amount"]:,.0f}
    Số tài khoản: 106000316181 - Vietinbank - Hoàng Việt Anh"""
            })
            
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    
    # Xuất dữ liệu ra Excel bằng pandas + openpyxl
    df = pd.DataFrame(all_records)
    df['notify'] = df.apply(lambda row: row['notify'] if row['due_amount'] not in [0,'0'] else '', axis = 1) # bỏ notify các dòng đẫ thanh toán
    df['zalo_link'] = df.apply(lambda row: row['zalo_link'] if row['due_amount'] not in [0,'0'] else '', axis = 1) # bỏ zalo_link các dòng đẫ thanh toán
    df = df[df['status'] != 'available'] # bỏ bớt các dòng chưa cho thuê
    col_remain = list(df.columns) 
    col_remain.remove('status')
    df = df[col_remain]
    df.sort_values(by = ['month','room'], ascending = [False, True], inplace = True)
    df.to_excel(file_report, index=False, sheet_name="Report", engine="openpyxl")
    # Mở lại file bằng openpyxl để chỉnh sửa
    wb = load_workbook(file_report)
    ws = wb["Report"]
    # Auto-fit độ rộng cột
    for col in ws.columns:
        max_len = 0
        min_len = 55
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 2 if max_len < min_len else min_len # để cột cuối ko bị quá dài
    # Định dạng tất cả cột số thành có phân cách hàng nghìn
    for col in ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in col:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"

    zalo_col = None
    name_col = None
    room_col = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == "zalo_link":
            zalo_col = col
        elif ws.cell(row=1, column=col).value == "name":
            name_col = col
        elif ws.cell(row=1, column=col).value == "room":
            room_col = col
    for row in range(2, ws.max_row + 1):
        url = ws.cell(row=row, column=zalo_col).value
        name = ws.cell(row=row, column=name_col).value
        room = ws.cell(row=row, column=room_col).value
        if url and str(url).startswith("http"):
            ws.cell(row=row, column=zalo_col).hyperlink = url
            ws.cell(row=row, column=zalo_col).value = f"Zalo {room} {name}"
            ws.cell(row=row, column=zalo_col).style = "Hyperlink"
    ws.freeze_panes = "A2"
    wb.save(file_report)
    
    with open(file_room, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)
    with open(file_price, "w", encoding="utf-8") as f:
        json.dump(price, f, ensure_ascii=False, indent=4)
    with open(file_tenant, "w", encoding="utf-8") as f:
        json.dump(tenant, f, ensure_ascii=False, indent=4)
    with open(file_account, "w", encoding="utf-8") as f:
        json.dump(account, f, ensure_ascii=False, indent=4)

    print("✅ created rent_report.xlsx,room.json,price.json,tenant.json,account.json")

# =============================================================================
# gen link webpage
# =============================================================================
def view():
    from IPython.display import HTML
    url = "https://sites.google.com/view/trosupham2vietxocodien"
    return(HTML(f'<a href="{url}" target="_blank">👉 Mở trang web</a>'))

# =============================================================================
# trả json các bảng rooms, tenants, prices
# =============================================================================
def query(table):
    import sqlite3
    import json
    safe_mount_drive()
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    try:
        cursor.execute(f"SELECT * FROM {table} WHERE id = 1")
        x = json.loads(cursor.fetchone()[1])
    except Exception as ex:
        print(ex)
    finally:
        conn.close()
    return(x)

# =============================================================================
# trả df các bảng cashflow, tong_diennuoc, chikhac
# =============================================================================
def querydf(table):
    safe_mount_drive()
    import pandas as pd
    import sqlite3
    conn = sqlite3.connect(db_file)
    df = pd.read_sql_query(f"SELECT * FROM {table}", conn)
    conn.close()
    return df

# =============================================================================
# ('prices', '202507.R3.electric_price', 3000/"abc")
# - cập nhật thông tin các bảng theo đường dẫn
# =============================================================================
def update(table, object_address, value_update):
    """
    ('prices', '202507.R3.electric_price', 3000/"abc")
    """
    import sqlite3
    safe_mount_drive()
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    try:
        cursor.execute(f"""
                    UPDATE {table}
                    SET data = json_set(data, '$.{object_address}', ?)
                    WHERE id = 1
                    """, (value_update,)
            )
        conn.commit()
    except Exception as ex:
        print(ex)
    finally:
        conn.close()
# =============================================================================
# creating db file by import direct from json: price, room, tenant. auto delete if exists
# =============================================================================
def import_json():
    import json
    import sqlite3
    from google.colab import drive
    safe_mount_drive()
    with open(file_price) as file:
        price = json.loads(file.read())
    with open(file_room, "r") as f:
        room = json.loads(f.read())
    with open(file_tenant, "r") as f:
        tenant = json.loads(f.read())
    with open(file_account, "r") as f:
        accounts = json.loads(f.read())

    conn = sqlite3.connect("/content/drive/MyDrive/Dau_tu/data/inn.db")
    cursor = conn.cursor()

    cursor.execute("DROP TABLE IF EXISTS rooms")
    cursor.execute("DROP TABLE IF EXISTS prices")
    cursor.execute("DROP TABLE IF EXISTS tenants")
    cursor.execute("DROP TABLE IF EXISTS accounts")

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS rooms (
        id INTEGER PRIMARY KEY,
        data JSON
    )
    """)
    cursor.execute("INSERT INTO rooms (data) VALUES (?)", (json.dumps(room),))
    conn.commit()

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS prices (
        id INTEGER PRIMARY KEY,
        data JSON
    )
    """)
    cursor.execute("INSERT INTO prices (data) VALUES (?)", (json.dumps(price),))
    conn.commit()

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS tenants (
        id INTEGER PRIMARY KEY,
        data JSON
    )
    """)
    cursor.execute("INSERT INTO tenants (data) VALUES (?)", (json.dumps(tenant),))
    conn.commit()

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS accounts (
        id INTEGER PRIMARY KEY,
        data JSON
    )
    """)
    cursor.execute("INSERT INTO accounts (data) VALUES (?)", (json.dumps(accounts),))
    conn.commit()

    conn.close()

# =============================================================================
# insert water and electricity consumed
# =============================================================================
def diennuoc():
    """
    no input argument, insert water and electricity consumed, insert = null => not update
    """
    from datetime import datetime
    today = datetime.now()
    this_month = datetime.strftime(today, "%Y%m")
    mes_elec = f"Công tơ ĐIỆN tháng {this_month}: "
    mes_water = f"Công tơ NƯỚC tháng {this_month}: "
    room = input("Room: ").upper()
    rooms = query('rooms')[this_month]
    if room not in rooms:
        print("Room not valid")
        return
    elif rooms[room]['status'] != 'rented':
        print("Room not rented yet")
        return
    elec_end = input(mes_elec)
    elec_end = 0 if elec_end == '' else int(elec_end)
    water_end = input(mes_water)
    water_end = 0 if water_end == '' else int(water_end)
    elec_end = elec_end if (rooms[room]['electric_end'] is None or elec_end > rooms[room]['electric_end']) else rooms[room]['electric_end']
    water_end = water_end if (rooms[room]['water_end'] is None or water_end > rooms[room]['water_end']) else rooms[room]['water_end']
    update('rooms', f'{this_month}.{room}.electric_end', elec_end)
    update('rooms', f'{this_month}.{room}.water_end', water_end)
    print("done")

# =============================================================================
# add new room by insert data clob
# =============================================================================
def add_room(room_data):
    import json
    import sqlite3
    from datetime import datetime

    today = datetime.now()
    this_month = datetime.strftime(today, "%Y%m")
    
    month = input("Month: ")
    month = month if month != '' else this_month
    room_name = input("Room: ")
    if room_name not in ["R1", "R2", "R3", "R4", "R5", "R11", "R22", "R33", "R44", "R55"]:
        print(room_name, 'not in ["R1", "R2", "R3", "R4", "R5", "R11", "R22", "R33", "R44", "R55"]')
        return
    
    safe_mount_drive()
    conn = sqlite3.connect(db_file)
    sql = f"""
        UPDATE rooms
        SET data = json_set(data, '$.{month}.{room_name}', json(?))
        WHERE id = ?
    """
    conn.execute(sql, (json.dumps(room_data), 1))
    conn.commit()
    conn.close()
    
# =============================================================================
# add new room by insert data clob
# =============================================================================
def add_tenant(tenant_data):
    import json
    import sqlite3
    tenants = query('tenants')
    if tenant_data['phone'] != tenant_data['zalo']:
        key = input("chose phone or zalo because diff: 1/2")
        if key not in ['1', '2']:
            print("bad option")
            return
        key = tenant_data['phone'] if key == '1' else tenant_data['zalo']
    else:
        key = tenant_data['phone']
    if key in tenants:
        print(f"{key} already in tenants")
        return
    safe_mount_drive()
    conn = sqlite3.connect(db_file)
    sql = f"""
        UPDATE tenants
        SET data = json_set(data, '$.active.{key}', json(?))
        WHERE id = ?
    """
    conn.execute(sql, (json.dumps(tenant_data), 1))
    conn.commit()
    conn.close()
    print(key, 'added')
    print(query('tenants')[key])

# =============================================================================
# create inform every new month from previous one    
# =============================================================================
def new_month():
    import copy
    from datetime import datetime
    from dateutil.relativedelta import relativedelta
    data = query("rooms")
    last_month = max(data.keys())   
    new_month = datetime.strftime(datetime.now(), "%Y%m")
    
    if last_month ==  new_month:
        print(new_month, "exists")
        return
    data[new_month] = copy.deepcopy(data[last_month])
    
    for room, info in data[new_month].items():
        if room in ["R1", "R2", "R3", "R4", "R5"]:
            print(f"====={room}=====")
            # if info['bill'] and info['prepayment'] and info['payment']:
            #     if info['bill'] <= info['prepayment']:
            #        info['prepayment'] = info['prepayment'] - info['bill'] + info['payment']
            #     else:
            #        info['prepayment'] = info['payment'] - info['due_amount'] if info['payment'] > info['due_amount'] else 0
            # elif info['prepayment'] is not None and info['payment'] is not None:
            #     info['prepayment'] += info['payment']
            # elif info['prepayment'] is None:
            #     info['prepayment'] = info['payment'] 
            this_prepayment = 0 if (new_month not in info['prepayment'] or info['prepayment'][new_month] == '' or info['prepayment'][new_month] is None) else info['prepayment'][new_month]
            
            try:
                info["start_date"] = datetime.strftime(datetime.strptime(info["start_date"], "%d/%m/%Y") + relativedelta(months=1) , "%d/%m/%Y") if info["start_date"] is not None else None
            except Exception as ex:
                print("start_date", ex, info["start_date"])
            try:
                info["end_date"] = datetime.strftime(datetime.strptime(info["end_date"], "%d/%m/%Y") + relativedelta(months=1),"%d/%m/%Y") if info["end_date"] is not None else None
            except Exception as ex:
                print("end_date", ex, info["end_date"])
            try:
                info["due_date"] = datetime.strftime(datetime.strptime(info["due_date"], "%d/%m/%Y") + relativedelta(months=1),"%d/%m/%Y") if info["due_date"] is not None else None
            except Exception as ex:
                print("due_date", ex, info["due_date"])
            if info["bill"] < this_prepayment + info["payment"]:
                due_amount = 0
            else:
                due_amount = info["bill"] - this_prepayment - info["payment"]
            info["electric_start"] = info["electric_end"]  
            info["electric_end"]   = None  
            info["electric_fee"]   = None
            info["water_start"]    = info["water_end"]  
            info["water_end"] = None  
            info["water_fee"] = None
            info["bill"]      = 0
            info["payment"] = 0 
            info["prepayment_this"] = 0 
            info["payment_date"] = None 
            info["due_amount"]   = due_amount
            info["due_amount"]   = 0
    safe_mount_drive()
    import json
    import sqlite3
    conn = sqlite3.connect(db_file)
    sql = """
        UPDATE rooms
        SET data = json_set(data, '$.' || ?, json(?))
        WHERE id = ?
    """
    conn.execute(sql, (new_month, json.dumps(data[new_month]), 1))
    conn.commit()
    conn.close()
    automap_tenant() # tự động rà soát cập nhật lại thông tin người thuê, chỉ lấy thông tin người thuê đang active. lấy đúng room người thuê
    
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    wb = load_workbook(file_hangthang)
    source_sheet = wb[datetime.strftime(datetime.now() - relativedelta(months = 1), '%m-%Y')]
    new_sheet = wb.copy_worksheet(source_sheet)
    new_sheet.title = datetime.strftime(datetime.now(),'%m-%Y')    
    row_max = max(new_sheet["C1"].value, new_sheet["D1"].value)
    new_sheet["C1"] = 10
    new_sheet["D1"] = 10
    new_sheet["B1"] = None
    new_sheet["B2"] = None
    new_sheet["B8"] = None
    new_sheet["B9"] = None
    new_sheet["B11"] = None
    new_sheet["B12"] = None
    new_sheet["B14"] = None
    new_sheet["B15"] = None
    new_sheet["C17"] = f"='{datetime.strftime(datetime.now() - relativedelta(months = 1), '%m-%Y')}'!B17"
    new_sheet["C18"] = f"='{datetime.strftime(datetime.now() - relativedelta(months = 1), '%m-%Y')}'!B18"
    for col in ['F','G','H','I','J','K','L']:
        for row in range(3,row_max+1):
            new_sheet[f"{col}{row}"] = None
            new_sheet[f"{col}{row}"].font = Font(strike=False)
    wb._sheets.remove(new_sheet)
    wb._sheets.insert(0, new_sheet)
    wb.save(file_hangthang)
    
    print(new_month, "initialized. Reset any room: ")
    room_reset = input().upper()
    if room_reset == '':
        return
    if room_reset not in ["R1", "R2", "R3", "R4", "R5"]:
        print(room_reset, 'not in ["R1", "R2", "R3", "R4", "R5"]')
        return
    reset_room(room_reset)
    
# =============================================================================
# reset info of room for fresh
# =============================================================================
def reset_room(*room_reset):
    if len(room_reset) == 0:
        room = input("Room to reset: ").upper()
    else: 
        room = room_reset[0]
    if room not in ["R1", "R2", "R3", "R4", "R5", "R11", "R22", "R33", "R44", "R55"]:
        print(room, 'not in ["R1", "R2", "R3", "R4", "R5", "R11", "R22", "R33", "R44", "R55"]')
        return
    from datetime import datetime
    this_month = datetime.strftime(datetime.now(), "%Y%m")
    data = query("rooms")[this_month][room]
    for info in data:
        if info == 'status':
            data[info] = 'available'
        elif info == 'electric_start':
            data[info] = data[info] if data['electric_end'] is None else data['electric_end']
        elif info == 'water_start':
            data[info] = data[info] if data['water_end'] is None else data['water_end']
        elif info == 'num':
            data[info] = 0
        else:
            data[info] = None
    safe_mount_drive()
    import sqlite3
    import json
    conn = sqlite3.connect(db_file)
    sql = f"""
        UPDATE rooms
        SET data = json_set(data, '$.{this_month}.{room}', json(?))
        WHERE id = ?
    """
    conn.execute(sql, (json.dumps(data), 1))
    conn.commit()
    conn.close()
    print(room, "already reset")
    
# =============================================================================
# automatically map tenants info to rooms, only this month
# Tự động cập nhật, chỉ áp dụng cho tháng này
# Kiểm tra đối chiếu ngày bắt đầu thuê của từng tenant
# Tự động đếm và cập nhật thông tin số người thuê cho từng phòng
# Chỉ lấy thông tin của người thuê là main hoặc người thuê là main cuối cùng nếu cả 2 cùng là main
# =============================================================================
def automap_tenant():
    from datetime import datetime
    this_month = datetime.strftime(datetime.now(), "%Y%m")
    rooms = query("rooms")[this_month]
    tenants = query("tenants")['active']
    for r in rooms:
        num = 0
        for t in tenants:
            tenant = tenants[t]
            tenant_start = datetime.strftime(datetime.strptime(tenant['start_date'], '%d/%m/%Y'), "%Y%m")
            if tenant['room'] == r and tenant['main'] == 1 and tenant_start <= this_month:
                update('rooms', f'{this_month}.{r}.rent_price', tenant['rent_price'])
                update('rooms', f'{this_month}.{r}.deposit', tenant['deposit'])
                update('rooms', f'{this_month}.{r}.deposit_date', tenant['deposit_date'])
                update('rooms', f'{this_month}.{r}.phone', tenant['phone'])
                update('rooms', f'{this_month}.{r}.zalo', tenant['zalo'])
                update('rooms', f'{this_month}.{r}.name', tenant['name'])
            else: 
                print(f"check room: {tenant['room']} == {r}: {tenant['room'] == r}")
                print(f"check main: {tenant['main']} == 1: {tenant['main'] == 1}")
                print(f"check month: {tenant_start} <= {this_month}: {tenant_start <= this_month}")
            if tenant['room'] == r:
                num += 1
                update('rooms', f'{this_month}.{r}.num', num)
                
# =============================================================================
# manual map tenant info to room, only this month           
# cập nhật thông tin người thuê vào phòng bằng cạch chọn người thuê và phòng cần cập nhật
# Người thuê chỉ được cập nhật vào đúng phòng thuê đã khai báo trong tenant và main phải =1  
# =============================================================================
def manualmap_tenant():
    from datetime import datetime
    this_month = datetime.strftime(datetime.now(), "%Y%m")
    rooms = query("rooms")[this_month]
    tenants = query("tenants")['active']
    for tenant in tenants.keys():
        print(tenant,tenants[tenant]['room'])
    t = input("tenant: ")
    if t not in tenants:
        print(t, "not in tenants")
        return
    tenant = tenants[t]
    print(rooms.keys())
    r = input("room: ").upper()
    if r not in rooms:
        print(r, "not in rooms")
        return
    num = 0
    tenant_start = datetime.strftime(datetime.strptime(tenant['start_date'], '%d/%m/%Y'), "%Y%m")
    if tenant['room'] == r and tenant['main'] == 1 and tenant_start <= this_month:
        update('rooms', f'{this_month}.{r}.rent_price', tenant['rent_price'])
        update('rooms', f'{this_month}.{r}.deposit', tenant['deposit'])
        update('rooms', f'{this_month}.{r}.deposit_date', tenant['deposit_date'])
        update('rooms', f'{this_month}.{r}.phone', tenant['phone'])
        update('rooms', f'{this_month}.{r}.zalo', tenant['zalo'])
        update('rooms', f'{this_month}.{r}.name', tenant['name'])
    else:
        print(f"check room: {tenant['room']} == {r}: {tenant['room'] == r}")
        print(f"check main: {tenant['main']} == 1: {tenant['main'] == 1}")
        print(f"check month: {tenant_start} <= {this_month}: {tenant_start <= this_month}")
                          
    if tenant['room'] == r:
        num+= 1
        update('rooms', f'{this_month}.{r}.num', num)         
        
# =============================================================================
# change tenant status: tenant = "sdt", status = "active/deactive" 
# - cập nhật theo input
# - nếu là active thì tự động xóa ngày end_date là ngày chấm dứt thuê, cập nhật ngày start_date là ngày thuê theo input
# - nếu là deactive thì tự động cập nhật ngày end_date là ngày chấm dứt thuê, ngày start_date giữ nguyên
# =============================================================================
def change_tenant_status(**kwargs):
    from datetime import datetime
    def status_tenant(new_data):
        import sqlite3
        import json
        safe_mount_drive()
        conn = sqlite3.connect(db_file)
        cursor = conn.cursor()
        cursor.execute("""
                       UPDATE tenants
                       SET data = ?
                       WHERE id = 1
                       """, (json.dumps(new_data),))
        conn.commit()
        print("updated")
        conn.close()
    tenants = query("tenants")
    active = tenants['active']
    deactive = tenants['deactive']
    if kwargs['tenant'] not in active and kwargs['tenant'] not in deactive:
        print(kwargs['tenant'], "not exists")
        return
    if kwargs['status'] == 'active':
        if kwargs['tenant'] in active:
            print('already active')
            return
        else:
            active[kwargs['tenant']] = active.get(kwargs['tenant'], deactive[kwargs['tenant']])
            start_date = input("start_date (dd/mm/yyyy): ")
            room = input("room: ").upper()
            rent_price = int(input("rent_price: "))
            mess_deposit = f"deposit / enter = {rent_price}: "
            deposit = input(mess_deposit)
            deposit = int(deposit) if deposit != '' else rent_price
            deposit_date = input("deposit_date (dd/mm/yyyy): ")
            try:
                datetime.strptime(start_date, '%d/%m/%Y')
                datetime.strptime(deposit_date, '%d/%m/%Y')
                active[kwargs['tenant']]['start_date'] = start_date
                active[kwargs['tenant']]['end_date'] = None
                active[kwargs['tenant']]['rent_price'] = rent_price
                active[kwargs['tenant']]['deposit'] = deposit
                active[kwargs['tenant']]['deposit_date'] = deposit_date
                active[kwargs['tenant']]['room'] = room
            except Exception as ex:
                print(ex)
                return
            deactive.pop(kwargs['tenant'])
            tenants['active'],tenants['deactive'] =  active, deactive
            status_tenant(tenants)
            print('activated')
    elif kwargs['status'] == 'deactive':
        if kwargs['tenant'] in deactive:
            print('already deactive')
            return
        else:
            deactive[kwargs['tenant']] = deactive.get(kwargs['tenant'], active[kwargs['tenant']])
            end_date = input("end_date (dd/mm/yyyy): ")
            try:
                datetime.strptime(end_date, '%d/%m/%Y')
                deactive[kwargs['tenant']]['end_date'] = end_date
            except Exception as ex:
                print(ex)
                return
            active.pop(kwargs['tenant'])
            tenants['active'],tenants['deactive'] =  active, deactive
            status_tenant(tenants)
            print('deactivated')
            message = f"need to reset room {active[kwargs['tenant']]['Room']} [yes]: "
            ask = input(message).upper()
            if ask == "YES":
                reset_room(active[kwargs['tenant']]['Room'])
    else:
        print("wrong status")

# =============================================================================
# Thống kê chi phí, doanh thu và tính toán số tiền tồn trên tài khoản
# =============================================================================
def doanhthu():
    safe_mount_drive()
    import pandas as pd
    import sqlite3
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Border, Side
    from datetime import datetime
    from dateutil.relativedelta import relativedelta

    rooms = query('rooms')
    tenants = query('tenants')
    thu = []

    # --- Thu tiền phòng ---
    for month in rooms:
        for room in rooms[month]:
            if rooms[month][room]['payment'] and rooms[month][room]['payment'] > 0:
                thu.append({
                    "Ngày": rooms[month][room]['payment_date'],
                    "Số tiền thu": rooms[month][room]['payment'],
                    "Nội dung": f"Phòng {room}, tháng {month}",
                    "Tiền điện": rooms[month][room]['electric_fee'],
                    "Tiền nước": rooms[month][room]['water_fee'],
                    "Month": month
                })

    # --- Thu tiền cọc ---
    for status in tenants:
        for tenant in tenants[status]:
            if tenants[status][tenant]['deposit'] != 0:
                thu.append({
                    "Ngày": tenants[status][tenant]['deposit_date'],
                    "Số tiền thu": tenants[status][tenant]['deposit'],
                    "Nội dung": f"Phòng {tenants[status][tenant]['room']} đặt cọc",
                    "Tiền điện": None,
                    "Tiền nước": None,
                    "Month": None
                })

    # --- Tạo DataFrame ---
    df = pd.DataFrame(thu)
    df = df[df['Số tiền thu'] != 0]

    # Thêm cột tháng điện/nước (lùi 1 tháng)
    df['month_water_el'] = df['Month'].apply(
        lambda x: datetime.strftime(datetime.strptime(str(x), "%Y%m") - relativedelta(months=1), "%Y%m")
                  if pd.notna(x) else x
    )

    df_diennuoc = querydf("tong_diennuoc")
    df_diennuoc['Ngày'] = df_diennuoc.apply(lambda row: datetime.strftime(datetime.strptime(row['Month'], "%Y%m"),"%d/%m/%Y"), axis = 1)
    df_diennuoc["Số tiền chi"] = -df_diennuoc["Tien_dien"] - df_diennuoc["Tien_nuoc"]
    df_diennuoc["Nội dung chi"] = df_diennuoc.apply(lambda row: f"Tiền điện: {row['Tien_dien']:,}, tiền nước: {row['Tien_nuoc']:,}", axis = 1)
    df_diennuoc = df_diennuoc[['Ngày','Số tiền chi','Nội dung chi', 'month_water_el']]
    
    df_chikhac = querydf("chikhac")
    df_chikhac['Số tiền chi'] = 0 - df_chikhac['sotien_chi']
    df_chikhac = df_chikhac.rename(columns = {'date':'Ngày',
                                'noidung_chi':"Nội dung chi"})
    df_chikhac.drop(columns = ['id', 'sotien_chi'], inplace=True)

    df_cash = pd.read_excel(file_cashbanoi)
    total_in = df['Số tiền thu'].sum()
    total_out = - df_diennuoc["Số tiền chi"].sum() - df_chikhac["Số tiền chi"].sum()
    new_record = [["Tổng thu", total_in],
                  ['Tổng chi', total_out],
                  ['Chênh lệnh thu chi', total_in - total_out],
                  ['Tiền mặt bà đang cầm hộ', df_cash['amount'].sum()],
                  ['Số dư tài khoản', total_in - total_out - df_cash['amount'].sum()]
                 ]
    
    df = pd.concat([df, df_diennuoc, df_chikhac], ignore_index=True)   
    # Sắp xếp theo ngày
    df['Ngày_dt'] = pd.to_datetime(df['Ngày'], format="%d/%m/%Y", errors="coerce")
    df = df.sort_values(by="Ngày_dt", ascending=False).drop(columns=["Ngày_dt"])
    
    #tính toán số tiền chênh điện nước
    df_grouped = df.groupby("month_water_el")[["Tiền điện", "Tiền nước", "Số tiền chi"]].sum().reset_index()
    df_grouped['Chênh lệch'] = df_grouped["Tiền điện"] +  df_grouped["Tiền nước"] + df_grouped["Số tiền chi"]
    df_grouped.sort_values(by = 'month_water_el', ascending = False, inplace = True)
    
    # thêm các dòng tổng
    df = pd.concat([pd.DataFrame(new_record, columns = ['Ngày', 'Số tiền thu']), df], ignore_index=True)
    # --- Xuất Excel ---
    with pd.ExcelWriter(file_cashflow, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="CashFlow")
        df_grouped.to_excel(writer, index=False, sheet_name="WaterElec")

    wb = load_workbook(file_cashflow)
    
    # Danh sách sheet cần xử lý
    for sheet_name in ["CashFlow", "WaterElec"]:
        if sheet_name not in wb.sheetnames:
            continue
        
        ws = wb[sheet_name]
    
        # Auto-fit độ rộng cột
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            width = min(max_len + 2, 55)
            ws.column_dimensions[get_column_letter(col[0].column)].width = width
    
        # Border đôi cho dòng cuối
        double_bottom = Border(bottom=Side(style="double"))
    
        if sheet_name == "CashFlow":
            money_cols = ["Số tiền thu", "Tiền điện", "Tiền nước", "Số tiền chi"]
        else:  # WaterElec
            money_cols = ["Tiền điện", "Tiền nước", "Số tiền chi", "Chênh lệch"]
    
        # Định dạng số cho các cột tiền
        headers = [cell.value for cell in ws[1]]
        for col_name in money_cols:
            if col_name in headers:
                col_idx = headers.index(col_name) + 1
                for col_cells in ws.iter_cols(min_row=2, max_row=ws.max_row,
                                              min_col=col_idx, max_col=col_idx):
                    for cell in col_cells:
                        # Ép kiểu float nếu có thể
                        try:
                            cell.value = float(cell.value)
                        except (TypeError, ValueError):
                            continue
                        cell.number_format = "#,##0"
    
        # Bôi đậm 5 dòng đầu, dòng cuối có border đôi
        max_row = ws.max_row
        for row in range(1, 6+1):
            for cell in ws[row]:
                cell.font = Font(bold=True)
                if row == 6:
                    cell.border = double_bottom
        ws.freeze_panes = "A2"
    
    wb.save(file_cashflow)
    # --- Ghi đè bảng cashflow trong SQLite ---
    conn = sqlite3.connect(db_file)
    df.to_sql("cashflow", conn, if_exists="replace", index=False)
    conn.close()
    print("Đã cập nhật cashflow")
    

# import sqlite3
# conn = sqlite3.connect(db_file)
# cursor = conn.cursor()
# cursor.execute("""
# CREATE TABLE IF NOT EXISTS tong_diennuoc (
#     Month TEXT PRIMARY KEY,
#     So_dien REAL,
#     Tien_dien REAL,
#     Gia_dien REAL,
#     So_nuoc REAL,
#     Tien_nuoc REAL,
#     Gia_nuoc REAL,
#     month_water_el TEXT
# )
# """)
# conn.commit()
# conn.close()
# =============================================================================
# Nhập số tiền chi điện nước
# tong_diennuoc("{this_month}", sodien, tien dien, sonuoc, tiennuoc)
# =============================================================================
def tong_diennuoc(month, so_dien, tien_dien, so_nuoc, tien_nuoc):
    safe_mount_drive()
    import sqlite3
    import math
    from datetime import datetime
    from dateutil.relativedelta import relativedelta
    import time
    today = datetime.now()
    this_month = datetime.strftime(today, "%Y%m")

    if month != this_month:
        print("Tháng không phải tháng hiện tại")
        ask = input("Có muốn tiếp tục không [yessss]: ").lower()
        if ask != 'yessss':
            return
    month_water_el = datetime.strftime(datetime.strptime(str(month), "%Y%m") - relativedelta(months=1), "%Y%m")

    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()

    # Tính giá bình quân
    gia_dien = math.ceil(tien_dien / so_dien / 100) * 100 if so_dien > 0 else 0
    gia_nuoc = math.ceil(tien_nuoc / so_nuoc / 100) * 100 if so_nuoc > 0 else 0
    electric_price = query('prices')[month]['electric_price']
    water_price = query('prices')[month]['water_price']
    if gia_dien > electric_price:
        update('prices', f'{month}.electric_price', gia_dien)
    if gia_nuoc > water_price:
        update('prices', f'{month}.water_price', gia_nuoc)

    cursor.execute("""
        INSERT INTO tong_diennuoc (Month, So_dien, Tien_dien, Gia_dien, So_nuoc, Tien_nuoc, Gia_nuoc, month_water_el)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(Month) DO UPDATE SET
            So_dien=excluded.So_dien,
            Tien_dien=excluded.Tien_dien,
            Gia_dien=excluded.Gia_dien,
            So_nuoc=excluded.So_nuoc,
            Tien_nuoc=excluded.Tien_nuoc,
            Gia_nuoc=excluded.Gia_nuoc,
            month_water_el = excluded.month_water_el
    """, (month, so_dien, tien_dien, gia_dien, so_nuoc, tien_nuoc, gia_nuoc, month_water_el))
    conn.commit()
    conn.close()
    if tien_dien > 0:
        add_trans('vietinbank', month, time.time(), {"amount": tien_dien,"date": datetime.strftime(today, "%d/%m/%Y"),"pay_for": "principal","pay_type": "debit","remark": f"Điện {month}: {so_dien} số","followed": "", "followed_id":""})
    if tien_nuoc > 0:
        add_trans('vietinbank', month, time.time(), {"amount": tien_nuoc,"date": datetime.strftime(today, "%d/%m/%Y"),"pay_for": "principal","pay_type": "debit","remark": f"Nước {month}: {so_nuoc} số","followed": "", "followed_id":""})
    print(f"Đã lưu tháng {month}: Giá điện {gia_dien:,} đ/kWh, Giá nước {gia_nuoc:,} đ/m³")
    tinhtien(1)
    print("Đã cập nhật giá vào room")
    doanhthu()

# =============================================================================
# adtrans vào bảng accounts và cập nhật vào bảng chikhac thông qua hàm chikhac1()
# =============================================================================
def chikhac():
    from datetime import datetime
    import time
    amount = int(input('amount: '))
    if amount == 0:
        print('amount = 0')
        return
    date = input('dd/mm/yyyy: ')
    date = date if date != '' else datetime.strftime(datetime.now(), '%d/%m/%Y')
    try:
        month = datetime.strftime(datetime.strptime(date, '%d/%m/%Y'), '%Y%m')
    except Exception as ex:
        print(ex)
        return
    month = datetime.strftime(datetime.strptime(date, '%d/%m/%Y'), '%Y%m')
    remark = input("remark: ")
    timeStamp = time.time()
    add_trans('vietinbank', month, timeStamp, {"amount": amount,"date": date,"pay_for": "principal","pay_type": "debit","remark": remark,"followed": "", "followed_id":""})
    chikhac1(date, remark, amount, str(timeStamp))
# =============================================================================
# Nhập số tiền chi ra khác
# chi_khac('dd/mm/yyyy', "Chuyển tiền sang thấu chi", 5000000, "")
# =============================================================================
def chikhac1(date, noidung_chi, sotien_chi, ghichu):
    safe_mount_drive()
    import sqlite3
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO chikhac (date, noidung_chi, sotien_chi, ghichu)
        VALUES (?, ?, ?, ?)
    """, (date, noidung_chi, sotien_chi, ghichu))
    conn.commit()
    conn.close()
    print(f"Đã lưu ngày {date}: chi {sotien_chi:,} - ghi chú: {ghichu}")
    doanhthu()

# import sqlite3
# conn = sqlite3.connect(db_file)
# cursor = conn.cursor()
# cursor.execute("""
# CREATE TABLE IF NOT EXISTS chikhac (
#     id INTEGER PRIMARY KEY AUTOINCREMENT,
#     date TEXT NOT NULL,
#     noidung_chi TEXT NOT NULL,
#     sotien_chi REAL NOT NULL,
#     ghichu TEXT
# )
# """)
# conn.commit()
# conn.close()

# =============================================================================
# Xóa 1 dòng trong bảng chikhac trong trường hợp nhập nhầm thông tin
# =============================================================================
def xoa_chi_khac(record_id):
    df = querydf('chikhac')
    trans_id = list(df[df['id']==record_id]['ghichu'])[0]
    import sqlite3
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    # Xóa theo id
    cursor.execute("DELETE FROM chikhac WHERE id = ?", (record_id,))
    conn.commit()
    conn.close()
    print(f"Đã xóa bản ghi id={record_id} trong bảng chikhac")
    reverse_transaction(trans_id)
    doanhthu()
# =============================================================================
# insert customer payment
# =============================================================================
def pay():
    from datetime import datetime
    import time
    from dateutil.relativedelta import relativedelta
    timeStamp = time.time()
    today = datetime.now()
    this_month = datetime.strftime(today, "%Y%m")
    room = input("Room: ").upper()
    rooms = query('rooms')[this_month]
    if room not in rooms:
        print("Room not valid")
        return
    elif rooms[room]['status'] != 'rented':
        print("Room not rented yet")
        return
    paid = rooms[room]['payment']
    bill = rooms[room]['bill']
    prepayment_this = rooms[room]['prepayment_this']
    if paid != 0:
        message = f"{room} already paid: {paid:,.0f}\n[y] to continue: "
        ask = input(message)
        if ask.upper() != "Y":
            return
    this_pay = int(input("Payment: "))
    payment = paid + this_pay
    if payment > bill:
        prepayment_this = prepayment_this + payment - bill
    update('rooms', f'{this_month}.{room}.prepayment_this', prepayment_this)
    next_month = today + relativedelta(months=1)
    while prepayment_this > 0:
        if datetime.strftime(next_month,"%Y%m") not in rooms[room]['prepayment'] or rooms[room]['prepayment'][datetime.strftime(next_month,"%Y%m")] == 0 or rooms[room]['prepayment'][datetime.strftime(next_month,"%Y%m")] is None:
            pre_add = rooms[room]['rent_price'] if rooms[room]['rent_price'] < prepayment_this else prepayment_this
            pre_in = pre_add
        else:
            pre_add = rooms[room]['rent_price'] - rooms[room]['prepayment'][datetime.strftime(next_month,"%Y%m")] if rooms[room]['prepayment'][datetime.strftime(next_month,"%Y%m")] < rooms[room]['rent_price'] else 0 
            pre_in = rooms[room]['prepayment'][datetime.strftime(next_month,"%Y%m")] + pre_add
        prepayment_this -= pre_add 
        update('rooms', f'{this_month}.{room}.prepayment.{datetime.strftime(next_month, "%Y%m")}', pre_in)
        next_month = next_month + relativedelta(months=1)
    update('rooms', f'{this_month}.{room}.payment', payment)
    update('rooms', f'{this_month}.{room}.payment_date', datetime.strftime(today, "%d/%m/%Y"))
    print(f"{room} marked paid {payment:,.0f} at {datetime.strftime(today, "%d/%m/%Y")}")
    add_trans('vietinbank', this_month, timeStamp, {"amount": this_pay,"date": datetime.strftime(today, "%d/%m/%Y"),"pay_for": "principal","pay_type": "credit","remark": f"{room} pay {this_month}","followed": "", "followed_id":""})
    tinhtien(1) # (1) to avoid asking month
    doanhthu()
# =============================================================================
# Hủy các giao dịch thanh toán tiền trọ của người trọ
# =============================================================================
def unpay():
    from datetime import datetime
    room = input("Room: ").upper()
    month = input("Month [yyyymm]: ")
    try:
        datetime.strptime(month, "%Y%m")
    except Exception as ex:
        print("month error:", ex)
        return
    rooms = query('rooms')[month]
    if room not in rooms:
        print("Room not valid")
        return
    elif rooms[room]['status'] != 'rented':
        print("Room not rented yet")
        return
    update('rooms', f'{month}.{room}.payment', 0)
    update('rooms', f'{month}.{room}.payment_date', "")
    trans = query('accounts')['active']['vietinbank']['transaction'][month]
    for tran in trans:
        if trans[tran]['remark'] == f"{room} pay {month}":
            reverse_transaction(tran)
    doanhthu()





# =============================================================================
# 
# =============================================================================
# =============================================================================
# PHẦN CÁC QUẢN LÝ CHI TIÊU
# =============================================================================
# =============================================================================
# 
# =============================================================================
def update_excel(sheet_name, col, value, row = None):
    from openpyxl import load_workbook

    # load workbook (không phá format)
    wb = load_workbook(file_hangthang)

    # chọn sheet
    ws = wb[sheet_name]

    if col in ['G','K'] and row is None:
        if col == "G":
            max_row = ws["C1"].value
            ws["C1"].value = max_row + 1
        elif col == 'K':
            max_row = ws["D1"].value
            ws["D1"].value = max_row + 1
        for row_r in range(1, int(max_row) + 1):
            if ws[f"{col}{row_r}"].value not in (None, ""):
                last_row = row_r
        row = last_row + 1
                   
    # chỉ thay đổi VALUE của cell
    if col in ['H','I','L'] and row == 2:
        ws[f"{col}{row}"].value = value
    elif col in ['B','H','I', 'L']:
        current_value = ws[f"{col}{row}"].value if ws[f"{col}{row}"].value is not None else 0
        ws[f"{col}{row}"].value = current_value + value
    else:
        ws[f"{col}{row}"].value = value
    
    # lưu lại file
    wb.save(file_hangthang)   
    return(row)

def brief_bid():
    print("""
        ====================================
            bid_luong(): tạo số dư ban đầu mỗi tháng cho quỹ lương
            bid_banoi(): ghi nhận khoản bà trả vào quỹ lương
            bid_thuong(): ghi nhận khoản Thưởng vào quỹ lương
        ====================================
            bid_chikhac(): ghi nhận các khoản chi việc chung từ Lương
            bid_thauchi_u(): ghi nhận thấu chi KHÔNG tài sản
            bid_thauchi_s(): ghi nhận thấu chi có tài sản
        ====================================
            bid_ungtien(): ghi nhận các khoản tạm ck về thấu chi để bớt gốc lãi
            bid_hoantien(): ghi nhận các khoản từ thấu chi về dda để hoàn lại đã ứng
        ====================================
            bid_tranomon(): ghi nhận giao dịch trả nợ món từ Trọ (v) hoặc BIDV
            bid_trathauchi_u(): ghi nhận giao dịch trả nợ thấu chi KHÔNG tài sản 
            bid_trathauchi_s(): ghi nhận giao dịch trả nợ thấu chi có tài sản
        ====================================
            reverse_transaction(trans_id): tìm và xóa tất cả các trans có trans_id trong tất cả các tài khoản
                delete_transaction(account, month, trans_id): xóa 01 giao dịch cho 1 tài khoản
        ====================================
                add_trans(account, month, timeStamp, *new_trans): tạo 01 giao dịch cho 1 tài khoản
        ====================================
          
          
""")
# =============================================================================
# Thêm các giao dịch cụ thể theo từng tài khoản
# =============================================================================
def add_trans(account, month, timeStamp, *new_trans):
    # today = datetime.now(ZoneInfo("Asia/Ho_Chi_Minh"))
    list_key = ['amount','date','pay_for','pay_type', 'followed','followed_id','remark']
    set_key = set(list_key)
    if len(new_trans) != 0:
        new_trans = new_trans[0]
        if set_key - set(new_trans.keys()) != set():
            print('thiếu',set_key - set(new_trans.keys()))
            return
    accounts = query('accounts')['active']
    acct = accounts[account]
    last_balance = acct['os_balance']
    account_type = acct['account_type']
    if account_type == 'dda':
        os_balance = last_balance + new_trans['amount'] if new_trans['pay_type'] == "credit" else last_balance - new_trans['amount']
    elif account_type in ['loan', 'overdraft']:
        os_balance = last_balance + new_trans['amount'] if new_trans['pay_type'] == "credit" else last_balance - new_trans['amount'] if new_trans['pay_for'] == 'principal' else last_balance
    new_trans['last_balance'] = last_balance
    new_trans['os_balance'] = os_balance
    update("accounts",f'active.{account}.os_balance', os_balance)
    patch = {
        "active": {
            account: {
                "transaction": {
                    month: {
                        timeStamp: new_trans
                    }
                }
            }
        }
    }
    import sqlite3
    import json
    safe_mount_drive()
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    try:
        cursor.execute("""
                            UPDATE accounts
                            SET data = json_patch(data, ?)
                            WHERE id = 1
                        """, (json.dumps(patch),))
        conn.commit()
    except Exception as ex:
        print(ex)
    finally:
        conn.close()
    
    if new_trans['followed'] == "yes":
        new_trans1 = new_trans.copy()
        new_trans1['account'] = account
        new_trans1['month'] = month
        patch = {
            "followed": {
                timeStamp: new_trans1
            }
        }
        import sqlite3
        import json
        safe_mount_drive()
        conn = sqlite3.connect(db_file)
        cursor = conn.cursor()
        try:
            cursor.execute("""
                                UPDATE accounts
                                SET data = json_patch(data, ?)
                                WHERE id = 1
                            """, (json.dumps(patch),))
            conn.commit()
        except Exception as ex:
            print(ex)
        finally:
            conn.close()
# =============================================================================
# xóa 01 bản ghi giao dịch
# =============================================================================
def delete_transaction(account, month, trans_id):
    import sqlite3
    safe_mount_drive()
    if account not in list(query("accounts")['active'].keys()) or month not in list(query("accounts")['active'][account]['transaction'].keys()) or trans_id not in list(query("accounts")['active'][account]['transaction'][month].keys()):
        print(account, month, trans_id, "data not found")
        return
    account_type = query("accounts")['active'][account]['account_type']
    os_balance = query("accounts")['active'][account]['os_balance']
    pay_type = query("accounts")['active'][account]['transaction'][month][trans_id]['pay_type']
    pay_for = query("accounts")['active'][account]['transaction'][month][trans_id]['pay_for']
    amount = query("accounts")['active'][account]['transaction'][month][trans_id]['amount']
    if account_type in ['loan', 'overdraft']:
        if pay_type == 'credit':
           os_balance -= amount
        elif pay_type == 'debit' and pay_for == 'principal':
            os_balance += amount 
    elif account_type == 'dda':
        if pay_type == 'credit':
            os_balance -= amount
        elif pay_type == 'debit':
            os_balance += amount 
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    try:
        # JSON path an toàn: quote các key
        path = f'$.active."{account}".transaction."{month}"."{trans_id}"'
        cursor.execute("""
            UPDATE accounts
            SET data = json_remove(data, ?)
            WHERE id = 1;
        """, (path,))
        conn.commit()
    except Exception as ex:
        print(ex)
    finally:
        conn.close()
    update("accounts",f'active.{account}.os_balance', os_balance)
# =============================================================================
# reserve lại 01 giao dịch bằng cách tìm và xóa các trans_id trong tất cả các tài khoản
# =============================================================================
def reverse_transaction(trans_id):
    import sqlite3
    accounts = query("accounts")['active']
    for account in list(accounts.keys()):
        for month in list(accounts[account]['transaction'].keys()):
            if trans_id in list(accounts[account]['transaction'][month].keys()):
                delete_transaction(account, month, trans_id)
    tran_ids = query("accounts")['followed']
    trans_id = str(trans_id)
    for id_tr in list(tran_ids.keys()):
        if id_tr == trans_id:
            conn = sqlite3.connect(db_file)
            cursor = conn.cursor()
            try:
                # JSON path an toàn: quote các key
                path = f'$.followed."{trans_id}"'
                cursor.execute("""
                    UPDATE accounts
                    SET data = json_remove(data, ?)
                    WHERE id = 1;
                """, (path,))
                conn.commit()
            except Exception as ex:
                print(ex)
            finally:
                conn.close()
                
# =============================================================================
# Tiêu thấu chi tín chấp 
# =============================================================================
def bid_thauchi_u():
    from datetime import datetime
    import time
    amount = int(input('amount: '))
    if amount == 0:
        print('amount = 0')
        return
    date = input('dd/mm/yyyy: ')
    date = date if date != '' else datetime.strftime(datetime.now(), '%d/%m/%Y')
    try:
        month = datetime.strftime(datetime.strptime(date, '%d/%m/%Y'), '%Y%m')
    except Exception as ex:
        print(ex)
        return
    month = datetime.strftime(datetime.strptime(date, '%d/%m/%Y'), '%Y%m')
    remark = input("remark: ")
    ask_follow = input("followed - Enter = yes / no")
    followed = 'yes' if ask_follow == '' else ''
    timeStamp = time.time()
    add_trans('overdraft_unsecured', month, timeStamp, {"amount": amount,"date": date,"pay_for": "principal","pay_type": "credit","remark": remark, "followed": followed, "followed_id":""})
    row_u = update_excel(f"{datetime.strftime(datetime.now(),'%m-%Y')}", "G", remark)
    update_excel(f"{datetime.strftime(datetime.now(),'%m-%Y')}", "H", -amount, row_u)
    update_excel(f"{datetime.strftime(datetime.now(),'%m-%Y')}", "F", date, row_u)
    update_excel(f"{datetime.strftime(datetime.now(),'%m-%Y')}", "H", f"=Sum(H3:H{row_u})", 2)
# =============================================================================
# Tiêu thấu có tài sản
# =============================================================================
def bid_thauchi_s():
    from datetime import datetime
    import time
    amount = int(input('amount: '))
    if amount == 0:
        print('amount = 0')
        return
    date = input('dd/mm/yyyy: ')
    date = date if date != '' else datetime.strftime(datetime.now(), '%d/%m/%Y')
    try:
        month = datetime.strftime(datetime.strptime(date, '%d/%m/%Y'), '%Y%m')
    except Exception as ex:
        print(ex)
        return
    month = datetime.strftime(datetime.strptime(date, '%d/%m/%Y'), '%Y%m')
    remark = input("remark: ")
    ask_follow = input("followed - Enter = yes / no")
    followed = 'yes' if ask_follow == '' else ''
    timeStamp = time.time()
    add_trans('overdraft_secured', month, timeStamp, {"amount": amount,"date": date,"pay_for": "principal","pay_type": "credit","remark": remark, "followed": followed, "followed_id":""})
# =============================================================================
# Trả nợ món 
# =============================================================================
def bid_tranomon():
    from datetime import datetime
    import time
    timeStamp = time.time()
    interest = input('interest: ')
    interest = int(interest) if interest != '' else 0
    principal = input('principal: ')
    principal = int(principal) if principal != '' else 0
    if interest + principal == 0:
        print("no payment")
        return
    date = input('dd/mm/yyyy: ')
    date = date if date != '' else datetime.strftime(datetime.now(), '%d/%m/%Y')
    try:
        month = datetime.strftime(datetime.strptime(date, '%d/%m/%Y'), '%Y%m')
    except Exception as ex:
        print(ex)
        return
    month = datetime.strftime(datetime.strptime(date, '%d/%m/%Y'), '%Y%m')
    source = input("[v] = vietinbank / Enter = bidv: ")
    remark = input("remark: ")
    if source == 'v':
        if principal != 0:
            add_trans('loan_45', month, timeStamp, {"amount": principal,"date": date,"pay_for": "principal","pay_type": "debit","remark": remark,"followed": "", "followed_id":""})
            add_trans('vietinbank', month, timeStamp, {"amount": principal,"date": date,"pay_for": "principal","pay_type": "debit","remark": remark,"followed": "", "followed_id":""})
            chikhac1(date, remark, principal, str(timeStamp))
            update_excel(f"{datetime.strftime(datetime.now(),'%m-%Y')}", "B", principal, 9)
            update_excel(f"{datetime.strftime(datetime.now(),'%m-%Y')}", "B", principal, 14)
        if interest != 0:
            timeStamp = time.time()
            add_trans('loan_45', month, timeStamp, {"amount": interest,"date": date,"pay_for": "interest","pay_type": "debit","remark": remark,"followed": "", "followed_id":""})
            add_trans('vietinbank', month, timeStamp, {"amount": interest,"date": date,"pay_for": "interest","pay_type": "debit","remark": remark,"followed": "", "followed_id":""})
            chikhac1(date, remark, interest, str(timeStamp))
    else:
        if principal != 0:
            add_trans('loan_45', month, timeStamp, {"amount": principal,"date": date,"pay_for": "principal","pay_type": "debit","remark": remark,"followed": "", "followed_id":""})
        if interest != 0:
            timeStamp = time.time()
            add_trans('loan_45', month, timeStamp, {"amount": interest,"date": date,"pay_for": "interest","pay_type": "debit","remark": remark,"followed": "", "followed_id":""})
    doanhthu()
# =============================================================================
# Trả nợ thấu chi tín 
# =============================================================================
def bid_trathauchi_u():
    from datetime import datetime
    import time
    timeStamp = time.time()
    interest = input('interest: ')
    interest = int(interest) if interest != '' else 0
    principal = input('principal: ')
    principal = int(principal) if principal != '' else 0
    if interest + principal == 0:
        print("no payment")
        return
    date = input('dd/mm/yyyy: ')
    date = date if date != '' else datetime.strftime(datetime.now(), '%d/%m/%Y')
    try:
        month = datetime.strftime(datetime.strptime(date, '%d/%m/%Y'), '%Y%m')
    except Exception as ex:
        print(ex)
        return
    month = datetime.strftime(datetime.strptime(date, '%d/%m/%Y'), '%Y%m')
    source = input("[v] = vietinbank / Enter = bidv: ")
    remark = input("remark: ")
    followed = query('accounts')['followed']
    followed_list = list(followed.keys())
    for key in followed_list:
        if followed[key]['account'] == 'overdraft_unsecured':
            print(followed_list.index(key), key, followed[key]['remark'], followed[key]['amount'])
    ask_id = input("id: ")
    if ask_id == '':
        followed_id = ''
    else:
        followed_id = followed_list[int(ask_id)]
    
    if source == 'v':
        if principal != 0:
            add_trans('overdraft_unsecured', month, timeStamp, {"amount": principal,"date": date,"pay_for": "principal","pay_type": "debit","remark": remark, "followed": "", "followed_id":followed_id})
            add_trans('vietinbank', month, timeStamp, {"amount": principal,"date": date,"pay_for": "principal","pay_type": "debit","remark": remark,"followed": "", "followed_id":""})
            chikhac1(date, remark, principal, str(timeStamp))
            update_excel(f"{datetime.strftime(datetime.now(),'%m-%Y')}", "B", principal, 9)
            update_excel(f"{datetime.strftime(datetime.now(),'%m-%Y')}", "B", principal, 15)
        if interest != 0:
            timeStamp = time.time()
            add_trans('overdraft_unsecured', month, timeStamp, {"amount": interest,"date": date,"pay_for": "interest","pay_type": "debit","remark": remark,"followed": "", "followed_id":""})
            add_trans('vietinbank', month, timeStamp, {"amount": interest,"date": date,"pay_for": "interest","pay_type": "debit","remark": remark,"followed": "", "followed_id":""})
            chikhac1(date, remark, interest, str(timeStamp))
    else:
        if principal != 0:
            add_trans('overdraft_unsecured', month, timeStamp, {"amount": principal,"date": date,"pay_for": "principal","pay_type": "debit","remark": remark, "followed": "", "followed_id":followed_id})
            row_u = update_excel(f"{datetime.strftime(datetime.now(),'%m-%Y')}", "G", remark)
            update_excel(f"{datetime.strftime(datetime.now(),'%m-%Y')}", "I", principal, row_u)
            update_excel(f"{datetime.strftime(datetime.now(),'%m-%Y')}", "F", date, row_u)
            update_excel(f"{datetime.strftime(datetime.now(),'%m-%Y')}", "I", f"=Sum(I3:I{row_u})", 2)
        if interest != 0:
            timeStamp = time.time()
            add_trans('overdraft_unsecured', month, timeStamp, {"amount": interest,"date": date,"pay_for": "interest","pay_type": "debit","remark": remark,"followed": "", "followed_id":""})
    doanhthu()
# =============================================================================
# Trả nợ thấu chi có tài 
# =============================================================================
def bid_trathauchi_s():
    from datetime import datetime
    import time
    timeStamp = time.time()
    interest = input('interest: ')
    interest = int(interest) if interest != '' else 0
    principal = input('principal: ')
    principal = int(principal) if principal != '' else 0
    if interest + principal == 0:
        print("no payment")
        return
    date = input('dd/mm/yyyy: ')
    date = date if date != '' else datetime.strftime(datetime.now(), '%d/%m/%Y')
    try:
        month = datetime.strftime(datetime.strptime(date, '%d/%m/%Y'), '%Y%m')
    except Exception as ex:
        print(ex)
        return
    month = datetime.strftime(datetime.strptime(date, '%d/%m/%Y'), '%Y%m')
    source = input("[v] = vietinbank / Enter = bidv: ")
    remark = input("remark: ")
    followed = query('accounts')['followed']
    followed_list = list(followed.keys())
    for key in followed_list:
        if followed[key]['account'] == 'overdraft_secured':
            print(followed_list.index(key), key, followed[key]['remark'], followed[key]['amount'])
    ask_id = input("id: ")
    if ask_id == '':
        followed_id = ''
    else:
        followed_id = followed_list[int(ask_id)]
        
    if source == 'v':
        if principal != 0:
            add_trans('overdraft_secured', month, timeStamp, {"amount": principal,"date": date,"pay_for": "principal","pay_type": "debit","remark": remark, "followed": "", "followed_id":followed_id})
            add_trans('vietinbank', month, timeStamp, {"amount": principal,"date": date,"pay_for": "principal","pay_type": "debit","remark": remark,"followed": "", "followed_id":""})
            chikhac1(date, remark, principal, str(timeStamp))
        if interest != 0:
            timeStamp = time.time()
            add_trans('overdraft_secured', month, timeStamp, {"amount": interest,"date": date,"pay_for": "interest","pay_type": "debit","remark": remark,"followed": "", "followed_id":""})
            add_trans('vietinbank', month, timeStamp, {"amount": interest,"date": date,"pay_for": "interest","pay_type": "debit","remark": remark,"followed": "", "followed_id":""})
            chikhac1(date, remark, interest, str(timeStamp))
    else:
        if principal != 0:
            add_trans('overdraft_secured', month, timeStamp, {"amount": principal,"date": date,"pay_for": "principal","pay_type": "debit","remark": remark, "followed": "", "followed_id":followed_id})
        if interest != 0:
            timeStamp = time.time()
            add_trans('overdraft_secured', month, timeStamp, {"amount": interest,"date": date,"pay_for": "interest","pay_type": "debit","remark": remark,"followed": "", "followed_id":""})
    doanhthu() 
  
# =============================================================================
# giao dịch chi tiêu khác từ lương cố định trả nợ
# =============================================================================
def bid_chikhac():
    from datetime import datetime
    import time
    amount = int(input('amount: '))
    if amount == 0:
        print('amount = 0')
        return
    date = input('dd/mm/yyyy: ')
    date = date if date != '' else datetime.strftime(datetime.now(), '%d/%m/%Y')
    try:
        month = datetime.strftime(datetime.strptime(date, '%d/%m/%Y'), '%Y%m')
    except Exception as ex:
        print(ex)
        return
    month = datetime.strftime(datetime.strptime(date, '%d/%m/%Y'), '%Y%m')
    remark = 'chichung_' + input("remark: ")
    timeStamp = time.time()
    add_trans('bidv', month, timeStamp, {"amount": amount,"date": date,"pay_for": "principal","pay_type": "debit","remark": remark,"followed": "", "followed_id":""})
    row_u = update_excel(f"{datetime.strftime(datetime.now(),'%m-%Y')}", "K", remark)
    update_excel(f"{datetime.strftime(datetime.now(),'%m-%Y')}", "L", -amount, row_u)
    update_excel(f"{datetime.strftime(datetime.now(),'%m-%Y')}", "J", date, row_u)
    update_excel(f"{datetime.strftime(datetime.now(),'%m-%Y')}", "L", f"=Sum(L3:L{row_u})", 2)
# =============================================================================
# ứng tiền từ dda vào thấu chi 
# =============================================================================
def bid_ungtien():
    from datetime import datetime
    import time
    amount = int(input('amount: '))
    if amount == 0:
        print('amount = 0')
        return
    date = input('dd/mm/yyyy: ')
    date = date if date != '' else datetime.strftime(datetime.now(), '%d/%m/%Y')
    try:
        month = datetime.strftime(datetime.strptime(date, '%d/%m/%Y'), '%Y%m')
    except Exception as ex:
        print(ex)
        return
    month = datetime.strftime(datetime.strptime(date, '%d/%m/%Y'), '%Y%m')
    ask = input("1.unsecured/2.secured: ")    
    overdraft = {"1":"overdraft_unsecured","2":"overdraft_secured"}[ask]
    remark = "ungtien_" + input("remark: ")
    followed = "yes"
    timeStamp = time.time()
    add_trans('bidv', month, timeStamp, {"amount": amount,"date": date,"pay_for": "principal","pay_type": "debit","remark": remark,"followed": '', "followed_id":""})
    add_trans(overdraft, month, timeStamp, {"amount": amount,"date": date,"pay_for": "principal","pay_type": "debit","remark": remark,"followed": followed, "followed_id":""})
    row_u = update_excel(f"{datetime.strftime(datetime.now(),'%m-%Y')}", "G", remark)
    update_excel(f"{datetime.strftime(datetime.now(),'%m-%Y')}", "I", amount, row_u)
    update_excel(f"{datetime.strftime(datetime.now(),'%m-%Y')}", "F", date, row_u)
    update_excel(f"{datetime.strftime(datetime.now(),'%m-%Y')}", "I", f"=Sum(I3:I{row_u})", 2)
    
# =============================================================================
# hoàn lương từ thấu chi vào bidv qua vietinbank
# =============================================================================
def bid_hoantien():
    from datetime import datetime
    import time
    amount = int(input('amount: '))
    if amount == 0:
        print('amount = 0')
        return
    date = input('dd/mm/yyyy: ')
    date = date if date != '' else datetime.strftime(datetime.now(), '%d/%m/%Y')
    try:
        month = datetime.strftime(datetime.strptime(date, '%d/%m/%Y'), '%Y%m')
    except Exception as ex:
        print(ex)
        return
    month = datetime.strftime(datetime.strptime(date, '%d/%m/%Y'), '%Y%m')
    ask = input("1.from unsecured\2.from secured: ")    
    overdraft = {"1":"overdraft_unsecured","2":"overdraft_secured"}[ask]
    remark = "hoantien_" + overdraft
    followed = "yes"
    timeStamp = time.time()
    add_trans('bidv', month, timeStamp, {"amount": amount,"date": date,"pay_for": "principal","pay_type": "credit","remark": remark,"followed": '', "followed_id":""})
    add_trans(overdraft, month, timeStamp, {"amount": amount,"date": date,"pay_for": "principal","pay_type": "credit","remark": remark,"followed": followed, "followed_id":""})
    row_u = update_excel(f"{datetime.strftime(datetime.now(),'%m-%Y')}", "G", remark)
    update_excel(f"{datetime.strftime(datetime.now(),'%m-%Y')}", "H", -amount, row_u)
    update_excel(f"{datetime.strftime(datetime.now(),'%m-%Y')}", "F", date, row_u)
    update_excel(f"{datetime.strftime(datetime.now(),'%m-%Y')}", "H", f"=Sum(H3:H{row_u})", 2)
# =============================================================================
# ghi có các khoản lương vào BIDV
# =============================================================================
def bid_luong():
    from datetime import datetime
    import time
    amount = 20_000_000
    date = input('dd/mm/yyyy: ')
    date = date if date != '' else datetime.strftime(datetime.now(), '%d/%m/%Y')
    try:
        month = datetime.strftime(datetime.strptime(date, '%d/%m/%Y'), '%Y%m')
    except Exception as ex:
        print(ex)
        return
    month = datetime.strftime(datetime.strptime(date, '%d/%m/%Y'), '%Y%m')
    remark = "salary_" + input("remark: ")
    timeStamp = time.time()
    add_trans('bidv', month, timeStamp, {"amount": amount,"date": date,"pay_for": "principal","pay_type": "credit","remark": remark,"followed": "", "followed_id":""})
        
# =============================================================================
# ghi có các khoản thưởng vào BIDV
# =============================================================================
def bid_thuong():
    from datetime import datetime
    import time
    amount = int(input('thuong: '))
    if amount == 0:
        print('amount = 0')
        return
    date = input('dd/mm/yyyy: ')
    date = date if date != '' else datetime.strftime(datetime.now(), '%d/%m/%Y')
    try:
        month = datetime.strftime(datetime.strptime(date, '%d/%m/%Y'), '%Y%m')
    except Exception as ex:
        print(ex)
        return
    month = datetime.strftime(datetime.strptime(date, '%d/%m/%Y'), '%Y%m')
    remark = "bonus_" + input("remark: ")
    timeStamp = time.time()
    add_trans('bidv', month, timeStamp, {"amount": amount,"date": date,"pay_for": "principal","pay_type": "credit","remark": remark,"followed": "", "followed_id":""})
    row_u = update_excel(f"{datetime.strftime(datetime.now(),'%m-%Y')}", "G", 'Thưởng')
    update_excel(f"{datetime.strftime(datetime.now(),'%m-%Y')}", "I", amount, row_u)
    update_excel(f"{datetime.strftime(datetime.now(),'%m-%Y')}", "F", date, row_u)
    update_excel(f"{datetime.strftime(datetime.now(),'%m-%Y')}", "I", f"=Sum(I3:I{row_u})", 2)
    update_excel(f"{datetime.strftime(datetime.now(),'%m-%Y')}", "B", amount, 11)
# =============================================================================
# Bà nội trả
# =============================================================================
def bid_banoi():
    from datetime import datetime
    import time
    amount = int(input('amount: '))
    if amount == 0:
        print('amount = 0')
        return
    date = input('dd/mm/yyyy: ')
    date = date if date != '' else datetime.strftime(datetime.now(), '%d/%m/%Y')
    try:
        month = datetime.strftime(datetime.strptime(date, '%d/%m/%Y'), '%Y%m')
    except Exception as ex:
        print(ex)
        return
    month = datetime.strftime(datetime.strptime(date, '%d/%m/%Y'), '%Y%m')
    remark = "banoi_" + month
    timeStamp = time.time()
    add_trans('bidv', month, timeStamp, {"amount": amount,"date": date,"pay_for": "principal","pay_type": "credit","remark": remark,"followed": "", "followed_id":""})
    row_u = update_excel(f"{datetime.strftime(datetime.now(),'%m-%Y')}", "G", 'Bà trả')
    update_excel(f"{datetime.strftime(datetime.now(),'%m-%Y')}", "I", amount, row_u)
    update_excel(f"{datetime.strftime(datetime.now(),'%m-%Y')}", "F", date, row_u)
    update_excel(f"{datetime.strftime(datetime.now(),'%m-%Y')}", "I", f"=Sum(I3:I{row_u})", 2)
    update_excel(f"{datetime.strftime(datetime.now(),'%m-%Y')}", "B", amount, 12)
